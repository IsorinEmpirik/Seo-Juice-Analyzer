"""
Routes de l'application Flask
"""
from flask import Blueprint, render_template, request, jsonify, session, current_app, send_file
from werkzeug.utils import secure_filename
import os
import io
from pathlib import Path
import logging
import uuid

from app.parsers import ScreamingFrogParser, AhrefsParser, GSCParser, EmbeddingsParser, cosine_similarity
from app.analyzer import SEOJuiceAnalyzer, recalculate_pagerank
from app.utils import get_csv_preview, detect_column_mapping
from app.gsc import GSCClient
from app import database as db

bp = Blueprint('main', __name__)
logger = logging.getLogger(__name__)

# Stockage temporaire des résultats (en production, utiliser Redis ou une BDD)
analysis_results = {}
# Stockage temporaire des fichiers uploadés
uploaded_files_storage = {}

import random
from urllib.parse import urlparse


def extract_slug_as_anchor(url):
    """
    Extrait le slug d'une URL et le transforme en ancre lisible.
    Ex: /blog/mon-super-article/ -> "mon super article"
    """
    try:
        parsed = urlparse(url)
        path = parsed.path.strip('/')
        if not path:
            return None
        # Prendre le dernier segment du chemin
        slug = path.split('/')[-1]
        if not slug:
            # Si le dernier segment est vide, prendre l'avant-dernier
            segments = [s for s in path.split('/') if s]
            slug = segments[-1] if segments else None
        if not slug:
            return None
        # Nettoyer le slug : remplacer tirets/underscores par des espaces
        anchor = slug.replace('-', ' ').replace('_', ' ')
        # Supprimer les extensions de fichiers (.html, .php, etc.)
        if '.' in anchor:
            anchor = anchor.rsplit('.', 1)[0]
        return anchor.strip() if anchor.strip() else None
    except Exception:
        return None


def generate_link_recommendations(priority_urls, embeddings_data, sf_parser, gsc_data=None, brand_keywords=None, non_indexable_urls=None, source_directory=None, max_links_per_priority=50):
    """
    Génère des recommandations de liens internes vers les pages prioritaires
    basées sur la similarité sémantique des embeddings.

    Args:
        priority_urls: Liste des URLs prioritaires
        embeddings_data: Dictionnaire {url: embedding_vector}
        sf_parser: Parser Screaming Frog (pour les liens existants)
        gsc_data: Données GSC agrégées par URL (optionnel)
        brand_keywords: Liste des mots-clés marque à exclure (optionnel)
        non_indexable_urls: Set d'URLs non indexables à exclure des sources (canonisées, noindex, etc.)
        source_directory: Répertoire source pour filtrer les pages candidates (ex: "/blog/")
        max_links_per_priority: Nombre maximum de liens par page prioritaire (garde-fou serveur)

    Returns:
        Liste de recommandations de liens
    """
    recommendations = []
    brand_keywords = [kw.lower() for kw in (brand_keywords or [])]
    non_indexable_urls = non_indexable_urls or set()

    # Récupérer les liens existants par source
    existing_links_by_source = sf_parser.get_links_by_source()

    # Construire un ensemble des liens existants DANS LE CONTENU ET LE FIL D'ARIANE
    # On ignore les liens dans Navigation (menu) et Pied de page - ils ne comptent pas pour le maillage
    existing_content_links_set = set()
    # Contenu + En-tête (breadcrumb/fil d'Ariane) - on exclut seulement Navigation et Pied de page
    content_positions = ['content', 'contenu', 'body', 'en-tête', 'header']

    for source, links in existing_links_by_source.items():
        for link in links:
            link_position = str(link.get('link_position', '')).lower().strip()
            # Ne compter que les liens dans le contenu
            if any(pos in link_position for pos in content_positions):
                existing_content_links_set.add((source, link['destination']))

    logger.info(f"Liens existants dans le contenu: {len(existing_content_links_set)}")

    if non_indexable_urls:
        logger.info(f"Pages non indexables exclues des recommandations: {len(non_indexable_urls)}")

    if source_directory:
        logger.info(f"Filtre répertoire source actif: {source_directory}")

    # Pour chaque page prioritaire
    for priority_url in priority_urls:
        priority_embedding = embeddings_data.get(priority_url)
        if not priority_embedding:
            logger.warning(f"Pas d'embedding trouvé pour l'URL prioritaire: {priority_url}")
            continue

        # Récupérer les mots-clés GSC de la page prioritaire (pour les ancres)
        priority_keywords = []
        if gsc_data and priority_url in gsc_data:
            url_gsc = gsc_data[priority_url]
            # Filtrer les mots-clés marque et trier par clics
            for kw in url_gsc.get('keywords', []):
                query_lower = kw['query'].lower()
                is_brand = any(brand in query_lower for brand in brand_keywords) if brand_keywords else False
                if not is_brand and kw.get('clicks', 0) > 0:
                    priority_keywords.append(kw)

        # Trier par clics décroissants pour favoriser les meilleurs mots-clés
        priority_keywords.sort(key=lambda x: x.get('clicks', 0), reverse=True)
        max_keywords = min(len(priority_keywords), 10)  # Utiliser jusqu'à 10 mots-clés différents

        # Collecter TOUTES les pages candidates avec leur similarité
        candidates = []

        for source_url, source_embedding in embeddings_data.items():
            # Ignorer la page prioritaire elle-même
            if source_url == priority_url:
                continue

            # Exclure les pages non indexables (canonisées, noindex)
            if source_url in non_indexable_urls:
                continue

            # Filtrer par répertoire source si spécifié
            if source_directory:
                source_path = urlparse(source_url).path or '/'
                if not source_path.startswith(source_directory):
                    continue

            # Vérifier si un lien existe déjà DANS LE CONTENU
            if (source_url, priority_url) in existing_content_links_set:
                continue

            # Calculer la similarité cosinus
            similarity = cosine_similarity(priority_embedding, source_embedding)

            # Garder toutes les pages avec une similarité > 0 (le filtrage fin sera en JS)
            if similarity > 0:
                candidates.append({
                    'source_url': source_url,
                    'similarity': round(similarity, 4)
                })

        # Trier par similarité décroissante
        candidates.sort(key=lambda x: x['similarity'], reverse=True)

        # Limiter au max_links_per_priority (garde-fou serveur)
        candidates = candidates[:max_links_per_priority]

        # Assigner les ancres avec variation (cycler à travers les mots-clés)
        for i, candidate in enumerate(candidates):
            if priority_keywords and max_keywords > 0:
                # Cycler à travers les mots-clés pour maximiser la variation
                selected_kw = priority_keywords[i % max_keywords]
                suggested_anchor = selected_kw['query']
            else:
                # Fallback: extraire l'ancre du slug de l'URL cible
                suggested_anchor = extract_slug_as_anchor(priority_url) or ""

            recommendations.append({
                'source_url': candidate['source_url'],
                'target_url': priority_url,
                'similarity': candidate['similarity'],
                'suggested_anchor': suggested_anchor,
            })

    # Trier globalement par similarité décroissante
    recommendations.sort(key=lambda x: x['similarity'], reverse=True)

    logger.info(f"Recommandations générées: {len(recommendations)} pour {len(priority_urls)} pages prioritaires")

    return recommendations


def allowed_file(filename):
    """Vérifie si le fichier est un CSV"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'csv'


@bp.route('/')
def index():
    """Page d'accueil"""
    gsc_connected = session.get('gsc_connected', False)
    gsc_account_id = session.get('gsc_account_id')
    return render_template('index.html',
                         gsc_connected=gsc_connected,
                         gsc_account_id=gsc_account_id)


@bp.route('/algo')
def algo():
    """Page explicative de l'algorithme"""
    return render_template('algo.html')


@bp.route('/preview/<upload_id>')
def preview(upload_id):
    """Page de prévisualisation avec mapping des colonnes"""
    if upload_id not in uploaded_files_storage:
        return render_template('error.html', message="Fichiers introuvables"), 404

    try:
        file_paths = uploaded_files_storage[upload_id]

        # Prévisualiser les CSV
        sf_columns, sf_rows = get_csv_preview(file_paths['screaming_frog'], num_rows=5)
        ahrefs_columns, ahrefs_rows = get_csv_preview(file_paths['ahrefs'], num_rows=5)

        # Détecter automatiquement les colonnes pertinentes
        sf_mapping = detect_column_mapping(sf_columns, 'screaming_frog')
        ahrefs_mapping = detect_column_mapping(ahrefs_columns, 'ahrefs')

        preview_data = {
            'screaming_frog': {
                'columns': sf_columns,
                'preview': sf_rows,
                'detected_mapping': sf_mapping
            },
            'ahrefs': {
                'columns': ahrefs_columns,
                'preview': ahrefs_rows,
                'detected_mapping': ahrefs_mapping
            }
        }

        # Ajouter prévisualisation GSC si présent
        if file_paths.get('gsc'):
            gsc_columns, gsc_rows = get_csv_preview(file_paths['gsc'], num_rows=5)
            preview_data['gsc'] = {
                'columns': gsc_columns,
                'preview': gsc_rows,
                'brand_keywords': file_paths.get('brand_keywords', [])
            }

        return render_template('preview.html', preview_data=preview_data, upload_id=upload_id)

    except Exception as e:
        logger.error(f"Erreur lors de la prévisualisation: {e}", exc_info=True)
        return render_template('error.html', message=f"Erreur: {str(e)}"), 500


@bp.route('/upload-preview', methods=['POST'])
def upload_preview():
    """Upload des fichiers et prévisualisation pour le mapping des colonnes"""
    try:
        # Vérifier que les fichiers requis sont présents
        if 'screamingfrog' not in request.files or 'ahrefs' not in request.files or 'embeddings' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'Les trois fichiers CSV sont requis (Screaming Frog, Ahrefs, Embeddings)'
            }), 400

        sf_file = request.files['screamingfrog']
        ahrefs_file = request.files['ahrefs']
        embeddings_file = request.files['embeddings']

        # Vérifier que les fichiers ne sont pas vides
        if sf_file.filename == '' or ahrefs_file.filename == '' or embeddings_file.filename == '':
            return jsonify({
                'status': 'error',
                'message': 'Les fichiers ne peuvent pas être vides'
            }), 400

        # Vérifier que ce sont des CSV
        if not allowed_file(sf_file.filename) or not allowed_file(ahrefs_file.filename) or not allowed_file(embeddings_file.filename):
            return jsonify({
                'status': 'error',
                'message': 'Seuls les fichiers CSV sont acceptés'
            }), 400

        # Créer un ID unique pour cet upload
        upload_id = str(uuid.uuid4())

        # Sauvegarder les fichiers temporairement
        upload_folder = Path(current_app.config['UPLOAD_FOLDER'])
        upload_folder.mkdir(exist_ok=True)

        sf_path = upload_folder / f"{upload_id}_screaming_frog.csv"
        ahrefs_path = upload_folder / f"{upload_id}_ahrefs.csv"
        embeddings_path = upload_folder / f"{upload_id}_embeddings.csv"

        sf_file.save(str(sf_path))
        ahrefs_file.save(str(ahrefs_path))
        embeddings_file.save(str(embeddings_path))

        logger.info(f"Fichiers uploadés pour preview {upload_id}")

        # Stocker les chemins
        uploaded_files_storage[upload_id] = {
            'screaming_frog': str(sf_path),
            'ahrefs': str(ahrefs_path),
            'gsc': None,
            'brand_keywords': [],
            'embeddings': str(embeddings_path),
            'priority_urls': []
        }

        # Récupérer les mots-clés marque (commun CSV et OAuth)
        brand_keywords = request.form.get('brand_keywords', '')
        if brand_keywords:
            keywords_list = [kw.strip() for kw in brand_keywords.split('\n') if kw.strip()]
            uploaded_files_storage[upload_id]['brand_keywords'] = keywords_list
            logger.info(f"Mots-clés marque: {keywords_list}")

        # Gérer le fichier GSC (optionnel) - CSV ou OAuth
        gsc_oauth_property = request.form.get('gsc_oauth_property', '')
        if gsc_oauth_property:
            # Mode OAuth : stocker la propriété choisie
            uploaded_files_storage[upload_id]['gsc_oauth_property'] = gsc_oauth_property
            logger.info(f"GSC OAuth propriété sélectionnée: {gsc_oauth_property}")
        elif 'gsc' in request.files:
            gsc_file = request.files['gsc']
            if gsc_file.filename != '' and allowed_file(gsc_file.filename):
                gsc_path = upload_folder / f"{upload_id}_gsc.csv"
                gsc_file.save(str(gsc_path))
                uploaded_files_storage[upload_id]['gsc'] = str(gsc_path)
                logger.info(f"Fichier GSC uploadé pour {upload_id}")

        # Récupérer les URLs prioritaires (optionnel)
        priority_urls = request.form.get('priority_urls', '')
        if priority_urls:
            # Séparer par lignes et nettoyer
            urls_list = [url.strip() for url in priority_urls.split('\n') if url.strip()]
            uploaded_files_storage[upload_id]['priority_urls'] = urls_list
            logger.info(f"URLs prioritaires: {len(urls_list)} URLs")

        # Récupérer le répertoire source (optionnel)
        source_directory = request.form.get('source_directory', '').strip()
        if source_directory:
            # S'assurer que le répertoire commence et finit par /
            if not source_directory.startswith('/'):
                source_directory = '/' + source_directory
            if not source_directory.endswith('/'):
                source_directory = source_directory + '/'
            uploaded_files_storage[upload_id]['source_directory'] = source_directory
            logger.info(f"Répertoire source: {source_directory}")

        # Prévisualiser les CSV
        sf_columns, sf_rows = get_csv_preview(str(sf_path), num_rows=3)
        ahrefs_columns, ahrefs_rows = get_csv_preview(str(ahrefs_path), num_rows=3)

        # Détecter automatiquement les colonnes pertinentes
        sf_mapping = detect_column_mapping(sf_columns, 'screaming_frog')
        ahrefs_mapping = detect_column_mapping(ahrefs_columns, 'ahrefs')

        response_data = {
            'status': 'success',
            'upload_id': upload_id,
            'screaming_frog': {
                'columns': sf_columns,
                'preview': sf_rows,
                'detected_mapping': sf_mapping
            },
            'ahrefs': {
                'columns': ahrefs_columns,
                'preview': ahrefs_rows,
                'detected_mapping': ahrefs_mapping
            }
        }

        # Ajouter prévisualisation GSC si présent (CSV ou OAuth)
        if uploaded_files_storage[upload_id]['gsc']:
            gsc_columns, gsc_rows = get_csv_preview(uploaded_files_storage[upload_id]['gsc'], num_rows=3)
            response_data['gsc'] = {
                'source': 'csv',
                'columns': gsc_columns,
                'preview': gsc_rows,
                'brand_keywords': uploaded_files_storage[upload_id]['brand_keywords']
            }
        elif uploaded_files_storage[upload_id].get('gsc_oauth_property'):
            response_data['gsc'] = {
                'source': 'oauth',
                'property': uploaded_files_storage[upload_id]['gsc_oauth_property'],
                'brand_keywords': uploaded_files_storage[upload_id]['brand_keywords']
            }

        return jsonify(response_data)

    except Exception as e:
        logger.error(f"Erreur lors de la prévisualisation: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Erreur lors de la prévisualisation: {str(e)}'
        }), 500


@bp.route('/analyze', methods=['POST'])
def analyze():
    """Lancer l'analyse complète"""
    try:
        # Vérifier que les fichiers sont présents
        if 'screamingfrog' not in request.files or 'ahrefs' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'Les deux fichiers CSV sont requis'
            }), 400

        sf_file = request.files['screamingfrog']
        ahrefs_file = request.files['ahrefs']

        # Vérifier que les fichiers ne sont pas vides
        if sf_file.filename == '' or ahrefs_file.filename == '':
            return jsonify({
                'status': 'error',
                'message': 'Les fichiers ne peuvent pas être vides'
            }), 400

        # Vérifier que ce sont des CSV
        if not allowed_file(sf_file.filename) or not allowed_file(ahrefs_file.filename):
            return jsonify({
                'status': 'error',
                'message': 'Seuls les fichiers CSV sont acceptés'
            }), 400

        # Créer un ID unique pour cette analyse
        analysis_id = str(uuid.uuid4())

        # Sauvegarder les fichiers temporairement
        upload_folder = Path(current_app.config['UPLOAD_FOLDER'])
        upload_folder.mkdir(exist_ok=True)

        sf_path = upload_folder / f"{analysis_id}_screaming_frog.csv"
        ahrefs_path = upload_folder / f"{analysis_id}_ahrefs.csv"

        sf_file.save(str(sf_path))
        ahrefs_file.save(str(ahrefs_path))

        logger.info(f"Fichiers uploadés pour l'analyse {analysis_id}")

        # Récupérer la configuration
        config = {
            'backlink_score': int(request.form.get('backlink_score', 10)),
            'iterations': int(request.form.get('iterations', 3)),
            'transmission_rate': float(request.form.get('transmission_rate', 85)) / 100,
            'content_link_rate': float(request.form.get('content_rate', 90)) / 100,
        }
        config['navigation_link_rate'] = 1 - config['content_link_rate']

        logger.info(f"Configuration: {config}")

        # Parser les CSV
        logger.info("Parsing Screaming Frog...")
        sf_parser = ScreamingFrogParser(str(sf_path))
        sf_parser.parse()

        logger.info("Parsing Ahrefs...")
        ahrefs_parser = AhrefsParser(str(ahrefs_path))
        ahrefs_parser.parse()

        # Lancer l'analyse
        logger.info("Lancement de l'analyse...")
        analyzer = SEOJuiceAnalyzer(config=config)
        results = analyzer.analyze(sf_parser, ahrefs_parser)

        # Stocker les résultats en mémoire
        analysis_results[analysis_id] = results

        # Sauvegarder dans la base de données pour l'historique
        db.save_analysis(analysis_id, results)

        # Nettoyer les fichiers temporaires
        sf_path.unlink()
        ahrefs_path.unlink()

        logger.info(f"Analyse {analysis_id} terminée avec succès")

        return jsonify({
            'status': 'success',
            'message': 'Analyse terminée avec succès',
            'analysis_id': analysis_id
        })

    except Exception as e:
        logger.error(f"Erreur lors de l'analyse: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Erreur lors de l\'analyse: {str(e)}'
        }), 500


@bp.route('/results/<analysis_id>')
def results(analysis_id):
    """Page de résultats"""
    if analysis_id not in analysis_results:
        return render_template('error.html', message="Analyse introuvable"), 404

    full_results = analysis_results[analysis_id]

    # Filtrer les données privées volumineuses pour le rendu template (tojson)
    results_for_template = {k: v for k, v in full_results.items() if not k.startswith('_')}

    return render_template('results.html', results=results_for_template, analysis_id=analysis_id)


@bp.route('/api/results/<analysis_id>')
def api_results(analysis_id):
    """API pour récupérer les résultats en JSON"""
    if analysis_id not in analysis_results:
        return jsonify({'status': 'error', 'message': 'Analyse introuvable'}), 404

    # Filtrer les données privées volumineuses
    results_clean = {k: v for k, v in analysis_results[analysis_id].items() if not k.startswith('_')}

    return jsonify({
        'status': 'success',
        'results': results_clean
    })


@bp.route('/analyze-with-mapping', methods=['POST'])
def analyze_with_mapping():
    """Lancer l'analyse avec mapping personnalisé des colonnes"""
    try:
        # Récupérer les données JSON
        data = request.get_json()

        if not data:
            return jsonify({
                'status': 'error',
                'message': 'Données manquantes'
            }), 400

        upload_id = data.get('upload_id')
        sf_mapping = data.get('sf_mapping', {})
        ahrefs_mapping = data.get('ahrefs_mapping', {})

        if upload_id not in uploaded_files_storage:
            return jsonify({
                'status': 'error',
                'message': 'Fichiers introuvables'
            }), 404

        file_paths = uploaded_files_storage[upload_id]

        # Créer un ID pour l'analyse
        analysis_id = str(uuid.uuid4())

        logger.info(f"Analyse {analysis_id} avec mapping personnalisé")
        logger.info(f"SF mapping: {sf_mapping}")
        logger.info(f"Ahrefs mapping: {ahrefs_mapping}")

        # Parser les CSV avec les mappings personnalisés
        logger.info("Parsing Screaming Frog...")
        sf_parser = ScreamingFrogParser(file_paths['screaming_frog'])
        sf_parser.parse()

        logger.info("Parsing Ahrefs...")
        ahrefs_parser = AhrefsParser(file_paths['ahrefs'])
        ahrefs_parser.parse()

        # Parser GSC si présent (CSV ou OAuth)
        gsc_data = None
        brand_keywords = file_paths.get('brand_keywords', [])

        if file_paths.get('gsc'):
            # Mode CSV
            logger.info("Parsing GSC (CSV)...")
            gsc_parser = GSCParser(file_paths['gsc'], brand_keywords=brand_keywords)
            gsc_parser.parse()
            gsc_data = gsc_parser.get_aggregated_by_url()
            logger.info(f"GSC CSV: {len(gsc_data)} URLs avec données de position")

        elif file_paths.get('gsc_oauth_property'):
            # Mode OAuth
            gsc_oauth_property = file_paths['gsc_oauth_property']
            gsc_account_id = session.get('gsc_account_id')
            logger.info(f"Récupération GSC via OAuth pour {gsc_oauth_property}...")

            if not gsc_account_id:
                raise ValueError("Session GSC expirée. Veuillez reconnecter votre compte Google Search Console.")

            gsc_client = GSCClient(
                client_id=current_app.config.get('GOOGLE_CLIENT_ID', ''),
                client_secret=current_app.config.get('GOOGLE_CLIENT_SECRET', ''),
                redirect_uri=current_app.config.get('GOOGLE_REDIRECT_URI', 'http://localhost:5000/oauth/callback'),
            )
            token_data = gsc_client.load_token(gsc_account_id)
            if not token_data:
                raise ValueError("Token GSC introuvable. Veuillez reconnecter votre compte Google Search Console.")

            credentials = gsc_client.get_credentials(token_data)
            if not credentials:
                raise ValueError("Credentials GSC invalides ou expirés. Veuillez reconnecter votre compte.")

            gsc_data = gsc_client.fetch_data(credentials, gsc_oauth_property)

            # Filtrer les mots-clés marque si spécifiés
            if brand_keywords and gsc_data:
                brand_lower = [kw.lower() for kw in brand_keywords]
                for url_key in gsc_data:
                    filtered_kws = [
                        kw for kw in gsc_data[url_key]['keywords']
                        if not any(brand in kw['query'].lower() for brand in brand_lower)
                    ]
                    removed = len(gsc_data[url_key]['keywords']) - len(filtered_kws)
                    gsc_data[url_key]['keywords'] = filtered_kws
                    gsc_data[url_key]['queries_count'] = len(filtered_kws)
                    gsc_data[url_key]['total_clicks'] = sum(kw['clicks'] for kw in filtered_kws)
                    gsc_data[url_key]['total_impressions'] = sum(kw['impressions'] for kw in filtered_kws)

            if gsc_data:
                logger.info(f"GSC OAuth: {len(gsc_data)} URLs avec données de position")
            else:
                logger.warning(f"GSC OAuth: aucune donnée récupérée pour {gsc_oauth_property}")

        # Parser Embeddings (obligatoire - compatible Gemini et OpenAI)
        logger.info("Parsing Embeddings...")
        embeddings_parser = EmbeddingsParser(file_paths['embeddings'])
        embeddings_parser.parse()
        embeddings_data = embeddings_parser.get_embeddings_by_url()
        non_indexable_urls = embeddings_parser.get_non_indexable_urls()
        embeddings_stats = embeddings_parser.get_parse_stats()
        logger.info(
            f"Embeddings: {embeddings_stats['valid_embeddings']} URLs, "
            f"{embeddings_stats['dimensions']} dimensions, "
            f"fournisseur: {embeddings_stats['provider']}"
        )

        # Lancer l'analyse
        logger.info("Lancement de l'analyse...")
        analyzer = SEOJuiceAnalyzer()
        results = analyzer.analyze(sf_parser, ahrefs_parser, gsc_data=gsc_data)

        # Générer les recommandations de liens pour les pages prioritaires
        priority_urls = file_paths.get('priority_urls', [])
        source_directory = file_paths.get('source_directory', '')
        if priority_urls:
            logger.info(f"Génération recommandations pour {len(priority_urls)} pages prioritaires...")
            link_recommendations = generate_link_recommendations(
                priority_urls=priority_urls,
                embeddings_data=embeddings_data,
                sf_parser=sf_parser,
                gsc_data=gsc_data,
                brand_keywords=brand_keywords,
                non_indexable_urls=non_indexable_urls,
                source_directory=source_directory or None,
            )
            results['link_recommendations'] = link_recommendations
            results['has_priority_urls'] = True
            results['priority_urls'] = priority_urls
            results['source_directory'] = source_directory
            logger.info(f"Recommandations générées: {len(link_recommendations)} liens suggérés")
        else:
            results['has_priority_urls'] = False
            results['link_recommendations'] = []
            results['priority_urls'] = []
            results['source_directory'] = ''

        # Stocker les données brutes pour le recalcul PageRank et le graphe
        results['_internal_links'] = analyzer.internal_links
        results['_backlinks'] = analyzer.backlinks
        results['_url_scores_keys'] = list(analyzer.url_scores.keys())
        results['_main_domain'] = analyzer.main_domain
        results['_embeddings_data'] = embeddings_data
        results['embeddings_stats'] = embeddings_stats
        results['analysis_mode'] = 'manual'

        # Stocker les résultats en mémoire
        analysis_results[analysis_id] = results

        # Sauvegarder dans la base de données pour l'historique
        db.save_analysis(analysis_id, results)

        # Nettoyer les fichiers temporaires
        Path(file_paths['screaming_frog']).unlink()
        Path(file_paths['ahrefs']).unlink()
        Path(file_paths['embeddings']).unlink()
        if file_paths.get('gsc'):
            Path(file_paths['gsc']).unlink()

        # Supprimer de la liste
        del uploaded_files_storage[upload_id]

        logger.info(f"Analyse {analysis_id} terminée avec succès")

        return jsonify({
            'status': 'success',
            'message': 'Analyse terminée avec succès',
            'analysis_id': analysis_id
        })

    except Exception as e:
        logger.error(f"Erreur lors de l'analyse avec mapping: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Erreur lors de l\'analyse: {str(e)}'
        }), 500


@bp.route('/export-sheets/<analysis_id>', methods=['POST'])
def export_to_sheets(analysis_id):
    """Exporter vers Google Sheets"""
    if analysis_id not in analysis_results:
        return jsonify({'status': 'error', 'message': 'Analyse introuvable'}), 404

    # TODO: Implémenter l'export vers Google Sheets
    return jsonify({
        'status': 'success',
        'message': 'Export Google Sheets sera implémenté prochainement'
    })


# ==================== ENDPOINTS GRAPHE & RECALCUL ====================

@bp.route('/api/graph-data/<analysis_id>')
def api_graph_data(analysis_id):
    """Retourne les données du graphe (noeuds + arêtes) pour Cytoscape.js"""
    if analysis_id not in analysis_results:
        return jsonify({'status': 'error', 'message': 'Analyse introuvable'}), 404

    results = analysis_results[analysis_id]
    internal_links = results.get('_internal_links', {})
    main_domain = results.get('_main_domain', '')
    embeddings_data = results.get('_embeddings_data', {})

    # Construire un dict URL -> données pour accès rapide
    url_data_map = {}
    for u in results['urls']:
        url_data_map[u['url']] = u

    # Noeuds
    nodes = []
    for u in results['urls']:
        path = urlparse(u['url']).path or '/'
        segments = [s for s in path.split('/') if s]
        directory = segments[0] if segments else '/'

        nodes.append({
            'id': u['url'],
            'label': path if len(path) <= 40 else '/' + '/'.join(segments[-2:]) if len(segments) >= 2 else path,
            'seo_score': u['seo_score'],
            'category': u['category'],
            'directory': directory,
            'backlinks_count': u['backlinks_count'],
            'internal_links_received': u['internal_links_received'],
            'internal_links_sent': u['internal_links_sent'],
            'status_code': u['status_code']
        })

    # Arêtes
    edges = []
    edge_id = 0
    for source_url, links in internal_links.items():
        if source_url not in url_data_map:
            continue
        for link in links:
            dest = link['destination']
            if dest not in url_data_map:
                continue
            if dest == source_url:
                continue  # Exclure self-links
            if main_domain and urlparse(dest).netloc != main_domain:
                continue

            # Calculer la similarité sémantique si embeddings disponibles
            similarity = None
            if embeddings_data:
                src_emb = embeddings_data.get(source_url)
                dst_emb = embeddings_data.get(dest)
                if src_emb and dst_emb:
                    similarity = round(cosine_similarity(src_emb, dst_emb), 4)

            edges.append({
                'id': f'e{edge_id}',
                'source': source_url,
                'target': dest,
                'link_type': link.get('link_position', 'Contenu'),
                'anchor': link.get('anchor', ''),
                'similarity': similarity
            })
            edge_id += 1

    # Récupérer la liste des répertoires uniques pour les filtres
    directories = sorted(set(n['directory'] for n in nodes))

    return jsonify({
        'status': 'success',
        'nodes': nodes,
        'edges': edges,
        'directories': directories,
        'has_embeddings': bool(embeddings_data),
        'total_nodes': len(nodes),
        'total_edges': len(edges)
    })


@bp.route('/api/recalculate-pagerank/<analysis_id>', methods=['POST'])
def api_recalculate_pagerank(analysis_id):
    """Recalcule le PageRank avec des liens ajoutés/supprimés"""
    if analysis_id not in analysis_results:
        return jsonify({'status': 'error', 'message': 'Analyse introuvable'}), 404

    results = analysis_results[analysis_id]
    data = request.get_json()

    if not data:
        return jsonify({'status': 'error', 'message': 'Données manquantes'}), 400

    added_links = data.get('added_links', [])
    removed_links = data.get('removed_links', [])

    internal_links = results.get('_internal_links', {})
    backlinks = results.get('_backlinks', {})
    url_keys = results.get('_url_scores_keys', [])

    try:
        new_scores = recalculate_pagerank(
            url_scores_keys=url_keys,
            internal_links=internal_links,
            backlinks=backlinks,
            added_links=added_links,
            removed_links=removed_links
        )

        # Calculer les deltas par rapport aux scores originaux
        original_scores = {u['url']: u['seo_score'] for u in results['urls']}
        deltas = {}
        for url, new_score in new_scores.items():
            old_score = original_scores.get(url, 0)
            delta = round(new_score - old_score, 2)
            if delta != 0:
                deltas[url] = {
                    'old_score': old_score,
                    'new_score': new_score,
                    'delta': delta
                }

        return jsonify({
            'status': 'success',
            'scores': new_scores,
            'deltas': deltas,
            'modifications_count': len(added_links) + len(removed_links)
        })

    except Exception as e:
        logger.error(f"Erreur recalcul PageRank: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==================== ENDPOINTS HISTORIQUE ====================

@bp.route('/api/history/domains')
def api_history_domains():
    """Liste tous les domaines avec historique d'analyses."""
    try:
        domains = db.get_all_domains()
        return jsonify({
            'status': 'success',
            'domains': domains
        })
    except Exception as e:
        logger.error(f"Erreur API domains: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@bp.route('/api/history/analyses')
def api_history_analyses():
    """Liste les analyses historiques, filtrable par domaine."""
    try:
        domain = request.args.get('domain')
        limit = int(request.args.get('limit', 50))

        analyses = db.get_analyses_for_domain(domain, limit)
        return jsonify({
            'status': 'success',
            'analyses': analyses
        })
    except Exception as e:
        logger.error(f"Erreur API analyses: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@bp.route('/api/history/analysis/<analysis_id>')
def api_history_analysis_details(analysis_id):
    """Récupère les détails d'une analyse historique."""
    try:
        analysis = db.get_analysis_details(analysis_id)
        if not analysis:
            return jsonify({'status': 'error', 'message': 'Analyse introuvable'}), 404

        return jsonify({
            'status': 'success',
            'analysis': analysis
        })
    except Exception as e:
        logger.error(f"Erreur API analysis details: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@bp.route('/api/history/compare', methods=['POST'])
def api_history_compare():
    """Compare deux analyses."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': 'Données manquantes'}), 400

        current_id = data.get('current_id')
        previous_id = data.get('previous_id')

        if not current_id or not previous_id:
            return jsonify({
                'status': 'error',
                'message': 'Les IDs des deux analyses sont requis'
            }), 400

        comparison = db.compare_analyses(current_id, previous_id)

        if 'error' in comparison:
            return jsonify({'status': 'error', 'message': comparison['error']}), 400

        return jsonify({
            'status': 'success',
            'comparison': comparison
        })
    except Exception as e:
        logger.error(f"Erreur API compare: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@bp.route('/api/history/evolution/<domain>')
def api_history_evolution(domain):
    """Récupère l'évolution des métriques pour un domaine."""
    try:
        limit = int(request.args.get('limit', 10))
        evolution = db.get_domain_evolution(domain, limit)

        return jsonify({
            'status': 'success',
            'domain': domain,
            'evolution': evolution
        })
    except Exception as e:
        logger.error(f"Erreur API evolution: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==================== EXPORT EXCEL ====================

@bp.route('/api/export-xlsx/<analysis_id>')
def export_xlsx(analysis_id):
    """Exporte les recommandations de liens en fichier Excel formaté."""
    if analysis_id not in analysis_results:
        return jsonify({'status': 'error', 'message': 'Analyse introuvable'}), 404

    results = analysis_results[analysis_id]
    recommendations = results.get('link_recommendations', [])

    if not recommendations:
        return jsonify({'status': 'error', 'message': 'Aucune recommandation de liens'}), 404

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from datetime import datetime

        # Récupérer les filtres depuis les query params
        threshold = float(request.args.get('threshold', 0))
        max_links = int(request.args.get('max_links', 0))
        target_filter = request.args.get('target', '')

        # Trier par page source puis par similarité décroissante
        sorted_recs = sorted(recommendations, key=lambda r: (r['source_url'], -r['similarity']))

        # Appliquer les filtres
        links_count_by_target = {}
        filtered_recs = []
        for rec in sorted_recs:
            if target_filter and rec['target_url'] != target_filter:
                continue
            if threshold > 0 and rec['similarity'] < threshold:
                continue
            if max_links > 0:
                links_count_by_target[rec['target_url']] = links_count_by_target.get(rec['target_url'], 0)
                if links_count_by_target[rec['target_url']] >= max_links:
                    continue
                links_count_by_target[rec['target_url']] += 1
            filtered_recs.append(rec)

        # Re-trier par page source pour le groupement
        filtered_recs.sort(key=lambda r: (r['source_url'], -r['similarity']))

        wb = Workbook()
        ws = wb.active
        ws.title = "Liens a Ajouter"

        # === Styles ===
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2B7A78', end_color='2B7A78', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6'),
        )
        url_font = Font(name='Calibri', size=10, color='0563C1', underline='single')
        normal_font = Font(name='Calibri', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Couleurs pour alternance par groupe de page source
        group_colors = [
            'F8F9FA',  # gris très clair
            'E8F4F8',  # bleu très clair
        ]

        # Gradient de couleurs pour la similarité (du plus faible au plus fort)
        def similarity_fill(score):
            """Retourne une couleur de remplissage basée sur le score de similarité."""
            if score >= 0.95:
                return PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')  # vert très foncé
            elif score >= 0.90:
                return PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # vert foncé
            elif score >= 0.85:
                return PatternFill(start_color='43A047', end_color='43A047', fill_type='solid')  # vert
            elif score >= 0.80:
                return PatternFill(start_color='66BB6A', end_color='66BB6A', fill_type='solid')  # vert clair
            elif score >= 0.75:
                return PatternFill(start_color='A5D6A7', end_color='A5D6A7', fill_type='solid')  # vert très clair
            elif score >= 0.70:
                return PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # vert pâle
            elif score >= 0.60:
                return PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # jaune pâle
            else:
                return PatternFill(start_color='FFECB3', end_color='FFECB3', fill_type='solid')  # orange pâle

        def similarity_font(score):
            """Retourne la police adaptée au fond."""
            if score >= 0.85:
                return Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            else:
                return Font(name='Calibri', size=10, bold=True, color='212529')

        # === En-têtes ===
        headers = ['Page Source', 'Page Cible', 'Similarite', 'Ancre Suggeree', 'Statut']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Figer la première ligne
        ws.freeze_panes = 'A2'

        # Filtre automatique
        ws.auto_filter.ref = f'A1:E{len(filtered_recs) + 1}'

        # === Données ===
        current_source = None
        group_idx = 0

        for row_idx, rec in enumerate(filtered_recs, 2):
            source_url = rec['source_url']
            target_url = rec['target_url']
            similarity = rec['similarity']
            anchor = rec['suggested_anchor']

            # Changer la couleur de groupe quand la page source change
            if source_url != current_source:
                current_source = source_url
                group_idx += 1

            group_fill = PatternFill(
                start_color=group_colors[group_idx % 2],
                end_color=group_colors[group_idx % 2],
                fill_type='solid'
            )

            # Page Source
            cell = ws.cell(row=row_idx, column=1, value=source_url)
            cell.font = url_font
            cell.fill = group_fill
            cell.alignment = left_align
            cell.border = thin_border

            # Page Cible
            cell = ws.cell(row=row_idx, column=2, value=target_url)
            cell.font = url_font
            cell.fill = group_fill
            cell.alignment = left_align
            cell.border = thin_border

            # Similarité (avec gradient de couleur)
            cell = ws.cell(row=row_idx, column=3, value=similarity)
            cell.number_format = '0.00%'
            cell.fill = similarity_fill(similarity)
            cell.font = similarity_font(similarity)
            cell.alignment = center_align
            cell.border = thin_border

            # Ancre Suggérée
            cell = ws.cell(row=row_idx, column=4, value=anchor)
            cell.font = normal_font
            cell.fill = group_fill
            cell.alignment = left_align
            cell.border = thin_border

            # Statut (vide, sera rempli par l'utilisateur)
            cell = ws.cell(row=row_idx, column=5, value='')
            cell.font = normal_font
            cell.fill = group_fill
            cell.alignment = center_align
            cell.border = thin_border

        # === Data validation pour la colonne Statut ===
        if len(filtered_recs) > 0:
            dv = DataValidation(
                type='list',
                formula1='"Fait,A faire,En cours,Ignore"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.prompt = 'Choisissez le statut'
            dv.promptTitle = 'Statut'
            ws.add_data_validation(dv)
            dv.add(f'E2:E{len(filtered_recs) + 1}')

        # === Largeurs de colonnes ===
        ws.column_dimensions['A'].width = 55  # Page Source
        ws.column_dimensions['B'].width = 55  # Page Cible
        ws.column_dimensions['C'].width = 14  # Similarité
        ws.column_dimensions['D'].width = 30  # Ancre
        ws.column_dimensions['E'].width = 14  # Statut

        # Hauteur de la ligne d'en-tête
        ws.row_dimensions[1].height = 30

        # === Feuille résumé ===
        ws_summary = wb.create_sheet('Resume', 0)
        ws_summary.sheet_properties.tabColor = '2B7A78'

        summary_data = [
            ('Analyse de Maillage Interne - Recommandations de Liens', ''),
            ('', ''),
            ('Date d\'export', datetime.now().strftime('%d/%m/%Y %H:%M')),
            ('Domaine', results.get('_main_domain', '')),
            ('Nombre total de recommandations', len(filtered_recs)),
            ('Pages sources uniques', len(set(r['source_url'] for r in filtered_recs))),
            ('Pages cibles uniques', len(set(r['target_url'] for r in filtered_recs))),
            ('Similarite moyenne', f"{sum(r['similarity'] for r in filtered_recs) / len(filtered_recs):.2%}" if filtered_recs else 'N/A'),
        ]

        if results.get('source_directory'):
            summary_data.append(('Repertoire source', results['source_directory']))

        if results.get('embeddings_stats'):
            stats = results['embeddings_stats']
            summary_data.append(('Embeddings', f"{stats['valid_embeddings']} pages, {stats['dimensions']} dimensions ({stats['provider']})"))
            if stats.get('non_indexable_count', 0) > 0:
                summary_data.append(('Pages non indexables exclues', stats['non_indexable_count']))

        # Titre
        title_cell = ws_summary.cell(row=1, column=1, value=summary_data[0][0])
        title_cell.font = Font(name='Calibri', bold=True, size=16, color='2B7A78')
        ws_summary.merge_cells('A1:B1')

        for row_idx, (label, value) in enumerate(summary_data[2:], 3):
            label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
            label_cell.font = Font(name='Calibri', bold=True, size=11)
            value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
            value_cell.font = Font(name='Calibri', size=11)

        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 60

        # Légende des couleurs de similarité
        legend_start = len(summary_data) + 4
        ws_summary.cell(row=legend_start, column=1, value='Legende - Similarite semantique').font = Font(name='Calibri', bold=True, size=12, color='2B7A78')

        legend_items = [
            (0.95, '95%+', 'Tres forte'),
            (0.90, '90-95%', 'Forte'),
            (0.85, '85-90%', 'Bonne'),
            (0.80, '80-85%', 'Correcte'),
            (0.75, '75-80%', 'Moyenne'),
            (0.70, '70-75%', 'Faible'),
            (0.60, '60-70%', 'Tres faible'),
        ]

        for i, (score, label, desc) in enumerate(legend_items):
            row = legend_start + 1 + i
            cell = ws_summary.cell(row=row, column=1, value=label)
            cell.fill = similarity_fill(score)
            cell.font = similarity_font(score)
            cell.alignment = center_align
            cell.border = thin_border
            ws_summary.cell(row=row, column=2, value=desc).font = normal_font

        # Activer la feuille "Liens a Ajouter" par défaut
        wb.active = wb.sheetnames.index('Liens a Ajouter')

        # Sauvegarder en mémoire
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Générer le nom du fichier
        domain = results.get('_main_domain', 'export')
        if not domain:
            domain = 'export'
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"{domain}_liens-a-ajouter_{date_str}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        logger.error(f"Erreur export XLSX: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500
